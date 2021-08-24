import openpyxl


class HomePageData:
    test_HomePageData = [{"homeAirport": "Heathrow", "destination": "Barbados", "travelDate": "16", "childAge": "7"},
                         {"homeAirport": "Gatwick", "destination": "Mauritius", "travelDate": "19", "childAge": "5"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\stuga\\Documents\\Selenium-Python\\ba.xlsx")
        sheet = book.active
        sheet.cell(row=1, column=2)
        Dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=1, column=1).value == test_case_name:
                for j in range(2, sheet.max_row + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
