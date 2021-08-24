import openpyxl

book = openpyxl.load_workbook("C:\\Users\\stuga\\Documents\\Selenium-Python\\ba.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
Dict = {}

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=1, column=1).value == "Client":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
